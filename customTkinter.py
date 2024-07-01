import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
import pandas as pd
from openpyxl.utils import get_column_letter
import openpyxl
from openpyxl import load_workbook

from openpyxl.styles import Alignment, Font, numbers, PatternFill, Color
class CustomTkinterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel File Handling App")
        
        # Initialize attributes
        self.file_uploaded = False
        self.file_name = ""

        # Center the window
        self.center_window(500, 400)

        ctk.set_appearance_mode("Dark")  # Modes: "System" (standard), "Dark", "Light"
        ctk.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"

        self.frame = ctk.CTkFrame(master=root)
        self.frame.pack(pady=20, padx=60, fill="both", expand=True)

        self.label = ctk.CTkLabel(master=self.frame, text="Excel File Handling App", font=("Roboto", 24))
        self.label.pack(pady=12, padx=10)

        self.upload_button = ctk.CTkButton(master=self.frame, text="Upload File", corner_radius=32, command=self.upload_file)
        self.upload_button.pack(pady=12, padx=10)

        self.save_button = ctk.CTkButton(master=self.frame, text="Process & Save File", corner_radius=32, command=self.save_file)
        self.save_button.pack(pady=12, padx=10)
        
        self.file_label = ctk.CTkLabel(master=self.frame, text="No file uploaded", font=("Roboto", 14))
        self.file_label.pack(pady=12, padx=10)

        self.bsp_rate_name = ctk.CTkLabel(master=self.frame, text="BSP Rate Value:", font=("Roboto", 16))
        self.bsp_rate_name.pack(pady=12, padx=10)

        self.bsp_rate_value_text = ctk.CTkLabel(master=self.frame, text=":", font=("Roboto", 14))
        self.bsp_rate_value_text.pack(pady=12, padx=10)

    def center_window(self, width, height):
        # Get the screen width and height
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Calculate the x and y coordinates to center the window
        x = (screen_width / 2) - (width / 2)
        y = (screen_height / 2) - (height / 2)
        
        # Set the geometry of the window
        self.root.geometry(f'{width}x{height}+{int(x)}+{int(y)}')
    


    def row_finder(workbook,sheet,target_value,target_value2,start_row,end_row):
        for row in range(1, sheet.max_row + 1):
            
            cell_value = sheet[f'A{row}'].value
            if cell_value:
                cell_value = str(cell_value).strip().upper()  # Normalize by stripping whitespace and converting to uppercase
                if cell_value == target_value:
                    start_row = row + 1
                elif cell_value == target_value2:
                    end_row = row
            
            # Exit loop if both start_row and end_row are found
            if start_row == None and end_row == None:
                start_row = 0
                end_row = 0
            
            if start_row and end_row:
                break       
        return start_row, end_row


    
        
    def upload_file(self):
        global file_path
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_path:
            self.file_uploaded = True  # Set flag to True
            self.file_name = file_path.split('/')[-1]  # Extract the file name
            self.file_label.configure(text=f"File: {self.file_name}")  # Update the label with the file name
            messagebox.showinfo("File Selected", f"File uploaded: {file_path}")
            # Handle file processing here
         # Load the Excel file
            df = pd.read_excel(file_path, sheet_name='Account Transactions')

                # Find the cell with "BSP Rate - Average Monthly"
            bsp_rate_cell = df[df.apply(lambda row: row.astype(str).str.contains(r'\(BSP Rate - Average Monthly\)').any(), axis=1)]

            if not bsp_rate_cell.empty:
                    # Get the row and column index of "BSP Rate - Average Monthly"
                bsp_rate_row_index = bsp_rate_cell.index[0]
                bsp_rate_col_index = bsp_rate_cell.columns[bsp_rate_cell.iloc[0] == "(BSP Rate - Average Monthly)"][0]
                    # Find the nearest number below "BSP Rate - Average Monthly" in the same column
                bsp_rate_value = None
            for i in range(bsp_rate_row_index + 1, len(df)):
                value = df.at[i, bsp_rate_col_index]
                if pd.notna(value) and isinstance(value, (int, float)):
                    bsp_rate_value = value
                    break

            if bsp_rate_value is not None:
                self.bsp_rate_value_text.configure(text=f"{bsp_rate_value}")
                messagebox.showinfo("BSP Rate Found", f"BSP Rate value: {bsp_rate_value}")
                df = pd.read_excel(file_path, sheet_name='Account Transactions')  # Replace with your sheet name if different

                # Function to find the column index containing 'BSP RATE'
                def find_bsp_rate_column(dataframe):
                    for idx, column in enumerate(dataframe.columns):
                        if 'BSP RATE' in str(column).upper():
                            return idx
                        if dataframe[column].astype(str).str.contains('BSP RATE', case=False, na=False).any():
                            return idx
                    return None

                # Find the index of the column containing 'BSP RATE'
                bsp_rate_column_index_w = find_bsp_rate_column(df)

                if bsp_rate_column_index_w is not None:
                    messagebox.showinfo("INFO",f"The column containing 'BSP RATE' is at index: {bsp_rate_column_index_w}")
                    df = pd.read_excel(file_path, sheet_name='Account Transactions', header=4) 
                    bsp_rate_column_index = bsp_rate_column_index_w
                    if bsp_rate_column_index is not None:
                        bsp_rate_column_letter = get_column_letter(bsp_rate_column_index + 1)
                        messagebox.showinfo("SUCCESS",f"\nThe column containing 'BSP RATE' is at index: {bsp_rate_column_index}, letter: {bsp_rate_column_letter}")
                        
                        php_column_name = 'Debit (PHP)'  # Replace with the actual column name after inspection
                        usd_column_name = 'USD'

                        # Define the exchange rate from PHP to USD (e.g., 1 PHP = 0.018 USD)
                        usdollar = bsp_rate_value
                        exchange_rate = 1 / usdollar
                        df[usd_column_name] = df[php_column_name] * exchange_rate


                        wb = openpyxl.load_workbook(file_path, keep_links=True)
                        sheet = wb['Account Transactions']
                        if bsp_rate_column_letter:

                            usd_column_letter = get_column_letter(bsp_rate_column_index + 1)
                            sheet[f'{usd_column_letter}5'].alignment = Alignment(horizontal='center', vertical='center')
                            sheet[f'{usd_column_letter}5'].font = Font(name='Arial',bold = True) 
                            for idx, value in enumerate(df[usd_column_name], start=6):  # start=6 because data starts from row 6
                                cell = f'{usd_column_letter}{idx}'
                                sheet[cell] = value
                                sheet[cell].alignment =  Alignment(horizontal='center', vertical='center')
                                sheet[cell].font = Font(name='Arial',size=8)
                                sheet[cell].number_format = '_$* #,##0.00_);[Red]($* (#,##0.00);_$* "-"??_)'
                                sheet[cell].fill = PatternFill(fill_type=None) 
                            sheet[f'{usd_column_letter}6'].value = bsp_rate_value
                            cell_color = Color(rgb="0070C0")
                            sheet[f'{usd_column_letter}6'].font = Font(name='Arial', color = cell_color, size= 9, italic= True)
                            
                           
                            #---------------------------------> CREATION OF TABLES <----------------------------------
                            
                            BSP_RATE = bsp_rate_value
                            workbook = openpyxl.load_workbook(file_path)
                            sheet = workbook['Account Transactions']
                                                 
                            def row_finder(workbook,sheet,target_value,target_value2,start_row,end_row):
                                for row in range(1, sheet.max_row + 1):
                                    
                                    cell_value = sheet[f'A{row}'].value
                                    if cell_value:
                                        cell_value = str(cell_value).strip().upper()  # Normalize by stripping whitespace and converting to uppercase
                                        if cell_value == target_value:
                                            start_row = row + 1
                                        elif cell_value == target_value2:
                                            end_row = row
                                    
                                    # Exit loop if both start_row and end_row are found
                                    if start_row == None and end_row == None:
                                        start_row = 0
                                        end_row = 0
                                    
                                    if start_row and end_row:
                                        break       
                                return start_row, end_row
                            

                            start_row = None
                            end_row = None

                            consumable_value = row_finder(workbook,sheet,'CONSUMABLES','TOTAL CONSUMABLES',start_row,end_row)
                            pin_value = row_finder(workbook, sheet, 'COST OF RAW MATERIALS - PINS', 'TOTAL COST OF RAW MATERIALS - PINS', start_row, end_row)
                            vertical_head_value = row_finder(workbook, sheet, 'COST OF RAW MATERIALS - VERTICAL HEAD', 'TOTAL COST OF RAW MATERIALS - VERTICAL HEAD', start_row, end_row)
                            factory_value = row_finder(workbook, sheet,'FACTORY SUPPLIES', 'TOTAL FACTORY SUPPLIES', start_row, end_row)
                            freight_in1_value = row_finder(workbook, sheet, 'FREIGHT IN - CONSUMABLES AND TOOLING EXPENSE', 'TOTAL FREIGHT IN - CONSUMABLES AND TOOLING EXPENSE', start_row, end_row)
                            freight_in2_value = row_finder(workbook, sheet, 'FREIGHT IN - DIRECT MATERIALS', 'TOTAL FREIGHT IN - DIRECT MATERIALS', start_row, end_row)
                            outside_value = row_finder(workbook,sheet, 'OUTSIDE SERVICES/FABRICATION', 'TOTAL OUTSIDE SERVICES/FABRICATION', start_row, end_row)
                            tooling_value = row_finder(workbook,sheet, 'TOOLING EXPENSE','TOTAL TOOLING EXPENSE', start_row, end_row)








                            #Columns 
                            product_column_letter = None                                                                                           
                            target_value3 = 'PRODUCT'
                            max_search_rows = 10


                            # Iterate through all rows and columns



                            for row in sheet.iter_rows(max_row=max_search_rows):
                                for cell in row:
                                    if cell.value:
                                        cell_value = str(cell.value).strip().upper()  # Normalize by stripping whitespace and converting to uppercase
                                        if cell_value == target_value3:
                                            product_column_letter = openpyxl.utils.get_column_letter(cell.column)
                                            break
                                
                                
                                if product_column_letter:
                                    break
                                
                            start_row = start_row


                            if product_column_letter == "A":
                                product_column_variable = 1
                            elif product_column_letter == "B":
                                product_column_variable = 2
                            elif product_column_letter == "C":
                                product_column_variable = 3
                            elif product_column_letter == "D":
                                product_column_variable = 4
                            elif product_column_letter == "E":
                                product_column_variable =5
                            elif product_column_letter == "F":
                                product_column_variable = 6
                            elif product_column_letter == "H":
                                product_column_variable = 7
                            elif product_column_letter == "I":
                                product_column_variable = 8
                            elif product_column_letter == "J":
                                product_column_variable = 9
                            column_data = []
                            # Extract data from column I (column index 9, assuming 'Product' is in column I)
                            column_i_data = []



                            for row_num in range(outside_value[0], outside_value[1]):
                                cell_value = sheet.cell(row=row_num, column=product_column_variable+1).value  # Column I corresponds to index 9 (A=1, B=2, ..., I=9)
                                column_i_data.append((row_num, cell_value)) 


                            #THIS CODE FINDS THE COLUMN OF DEBIT 
                            debit_colum_row = None
                            target_value4 = 'Debit (PHP)'
                            max_search_rows = 10



                            for row in sheet.iter_rows(max_row=max_search_rows): #SEARCHING FOR THE DEBIT COLUMN
                                for cell in row:
                                    if cell.value == target_value4:
                                        debit_colum_row = cell
                                        break
                                if debit_colum_row:
                                    break

                            debit_column_letter = debit_colum_row.coordinate[0] #Determines the column of the DEBIT(PHP)

                            
                            def categorize_rows(sheet, consumable_value_range, product_column_variable):
                                """
                                Categorize rows based on the product descriptions found in a specified column of a given sheet.

                                Parameters:
                                sheet (object): The sheet object to read data from.
                                consumable_value_range (tuple): A tuple containing the start and end row numbers (inclusive).
                                product_column_variable (int): The column index (0-based) where the product descriptions are located.

                                Returns:
                                dict: A dictionary with categorized rows.
                                """

                                # Initialize arrays to hold categorized row numbers
                                fabrication1 = []
                                fabrication2 = []
                                fabrication3 = []
                                probe_card = []
                                probe_cardv = []

                                # Dictionary to map product descriptions to their corresponding arrays
                                product_mapping = {
                                    'Fabrication1 - PCB/Board Repair': fabrication1,
                                    'Fabrication2 - Test Sockets': fabrication2,
                                    'Fabrication3 - Mechanical/General': fabrication3,
                                    'Probecard - Vertical': probe_cardv,
                                    'Probecard - Cantilever': probe_card
                                }

                                # Iterate through the specified range of rows
                                for row in range(consumable_value_range[0], consumable_value_range[1]):  # <-------------------------------------------
                                    cell_value = sheet.cell(row=row, column=product_column_variable + 1).value
                                    print(f"Row {row}, Column {product_column_variable + 1}: {cell_value}")  # Debug statement

                                    if cell_value is None:
                                        continue  # Skip None values

                                    # Add the row number to the corresponding array
                                    if cell_value in product_mapping:
                                        product_mapping[cell_value].append(row)

                                return product_mapping



                            def getting_data(sheet, debit_column_letter, big_data, BSP_rate):
                                fab1_final_value = 0
                                probe_card_final_value = 0
                                probe_cardv_final_value = 0
                                fab2_final_value = 0
                                fab3_final_value = 0


                                def get_cell_value(sheet, cell_coordinate):
                                    return sheet[cell_coordinate].value
                                
                                probe_card = big_data.get('Probecard - Cantilever', [])
                                probe_cardv = big_data.get('Probecard - Vertical', [])
                                fabrication1 = big_data.get('Fabrication1 - PCB/Board Repair', [])
                                fabrication2 = big_data.get('Fabrication2 - Test Sockets', [])
                                fabrication3 = big_data.get('Fabrication3 - Mechanical/General', [])

                                for rows in probe_card:
                                    cell_coordinate = f"{debit_column_letter}{rows}"
                                    cell_value = get_cell_value(sheet, cell_coordinate)
                                    if cell_value is not None:
                                        probe_card_final_value += cell_value
                                
                                for rows in probe_cardv:
                                    cell_coordinate = f"{debit_column_letter}{rows}"
                                    cell_value = get_cell_value(sheet, cell_coordinate)
                                    if cell_value is not None:
                                        probe_cardv_final_value += cell_value

                                for rows in fabrication1:
                                    cell_coordinate = f"{debit_column_letter}{rows}"
                                    cell_value = get_cell_value(sheet, cell_coordinate)
                                    if cell_value is not None:
                                        fab1_final_value += cell_value

                                for rows in fabrication2:
                                    cell_coordinate = f"{debit_column_letter}{rows}"
                                    cell_value = get_cell_value(sheet, cell_coordinate)
                                    if cell_value is not None:
                                        fab2_final_value += cell_value

                                for rows in fabrication3:
                                    cell_coordinate = f"{debit_column_letter}{rows}"
                                    cell_value = get_cell_value(sheet, cell_coordinate)
                                    if cell_value is not None:
                                        fab3_final_value += cell_value

                                # Convert to float and format the final values
                                probe_card_final_value = float(probe_card_final_value)
                                probe_cardv_final_value = float(probe_cardv_final_value)
                                fab1_final_value = float(fab1_final_value)
                                fab2_final_value = float(fab2_final_value)
                                fab3_final_value = float(fab3_final_value)

                                probe_card_final_value = "{:.2f}".format(probe_card_final_value / BSP_RATE)
                                probe_cardv_final_value = "{:.2f}".format(probe_cardv_final_value / BSP_RATE)
                                fab1_final_value = "{:.2f}".format(fab1_final_value / BSP_RATE)
                                fab2_final_value = "{:.2f}".format(fab2_final_value / BSP_RATE)
                                fab3_final_value = "{:.2f}".format(fab3_final_value / BSP_RATE)

                                # Calculate total value
                                total_value = float(probe_card_final_value) + float(probe_cardv_final_value) + float(fab1_final_value) + float(fab2_final_value) + float(fab3_final_value)
                                total_value = "{:.2f}".format(total_value)

                                return probe_card_final_value, probe_cardv_final_value, fab1_final_value, fab2_final_value, fab3_final_value, total_value

                            consume_row = categorize_rows(sheet, consumable_value, product_column_variable)
                            consume_data = getting_data(sheet, debit_column_letter, consume_row, BSP_RATE)

                            pin_row = categorize_rows(sheet, pin_value, product_column_variable)
                            pin_data = getting_data(sheet, debit_column_letter, pin_row, BSP_RATE)

                            vertical_row = categorize_rows(sheet, vertical_head_value, product_column_variable)
                            vertical_data = getting_data(sheet, debit_column_letter, vertical_row, BSP_RATE)

                            factory_row = categorize_rows(sheet, factory_value, product_column_variable)
                            factory_data = getting_data(sheet, debit_column_letter, factory_row, BSP_RATE)

                            freight1_row = categorize_rows(sheet, freight_in1_value, product_column_variable)
                            freight1_data = getting_data(sheet, debit_column_letter, freight1_row, BSP_RATE)

                            freight2_row = categorize_rows(sheet, freight_in2_value, product_column_variable)
                            freight2_data = getting_data(sheet, debit_column_letter, freight2_row, BSP_RATE)

                            outside_row = categorize_rows(sheet, outside_value, product_column_variable)
                            outside_data = getting_data(sheet, debit_column_letter, outside_row, BSP_RATE)

                            tooling_row = categorize_rows(sheet, tooling_value, product_column_variable)
                            tooling_data = getting_data(sheet, debit_column_letter, tooling_row, BSP_RATE)

                            def replace_zero_values(data):
                                return ["-" if value == "0.00" else value for value in data]


                            data_variables = [
                                ("Consumables", consume_data),
                                ("Cost of Raw Materials - Pins", pin_data),
                                ("Cost of Raw Materials - Vertical head", vertical_data),
                                ("Factory Supplies", factory_data),
                                ("Freight In - Consumables And Tooling Expense", freight1_data),
                                ("Freight In - Direct Materials", freight2_data),
                                ("Outside Services/Fabrication", outside_data),
                                ("Tooling Expense", tooling_data)
                            ]


                            #ADDITIONAL OF TOTAL VALUES IN COLUMN
                            probe_card_value = 0
                            probe_card_value = float(consume_data[0]) + float(pin_data[0]) + float(vertical_data[0]) + float(factory_data [0]) + float(freight1_data[0]) + float(freight2_data[0]) + float(outside_data[0]) + float(tooling_data[0])
                            probe_cardv_value = float(consume_data[1]) + float(pin_data[1]) + float(vertical_data[1]) + float(factory_data [1]) + float(freight1_data[1]) + float(freight2_data[1]) + float(outside_data[1]) + float(tooling_data[1])
                            fabrication1_value = float(consume_data[2]) + float(pin_data[2]) + float(vertical_data[2]) + float(factory_data [2]) + float(freight1_data[2]) + float(freight2_data[2]) + float(outside_data[2]) + float(tooling_data[2])
                            fabrication2_value = float(consume_data[3]) + float(pin_data[3]) + float(vertical_data[3]) + float(factory_data [3]) + float(freight1_data[3]) + float(freight2_data[3]) + float(outside_data[3]) + float(tooling_data[3])
                            fabrication3_value = float(consume_data[4]) + float(pin_data[4]) + float(vertical_data[4]) + float(factory_data [4]) + float(freight1_data[4]) + float(freight2_data[4]) + float(outside_data[4]) + float(tooling_data[4])
                            total_total_value = float(consume_data[5]) + float(pin_data[5]) + float(vertical_data[5]) + float(factory_data [5]) + float(freight1_data[5]) + float(freight2_data[5]) + float(outside_data[5]) + float(tooling_data[5])


                            print("Consume Data: DATA",consume_data[1])
                            # Extract the variable names and data

                            variable_names = [var[0] for var in data_variables]
                            data_values = [replace_zero_values(var[1]) for var in data_variables]

                            # Create a DataFrame
                            df = pd.DataFrame(data_values, columns=[
                                "Probecard - Cantilever",
                                "Probecard - Vertical",
                                "Fabrication1 - PCB/Board Repair",
                                "Fabrication2 - Test Sockets",
                                "Fabrication3 - Mechanical/General",
                                "Total Value"
                            ])


                            # Add the variable names as a column
                            df.insert(0, "Variable Name", variable_names)


                            new_row = {
                                "Variable Name": 'Total Value',  # Assuming 'Total Value' is the name for the first column
                                "Probecard - Cantilever": probe_card_value,
                                "Probecard - Vertical": probe_cardv_value,
                                "Fabrication1 - PCB/Board Repair": fabrication1_value,
                                "Fabrication2 - Test Sockets": fabrication2_value,
                                "Fabrication3 - Mechanical/General": fabrication3_value,
                                "Total Value": total_total_value
                            }

                            # Append the new row to the DataFrame
                            new_row_df = pd.DataFrame([new_row])

                            # Concatenate the original DataFrame with the new row DataFrame
                            df = pd.concat([df, new_row_df], ignore_index=True)


                            messagebox.showwarning("PANGET MO")
                            # Concatenate the original DataFrame with the new row DataFrame

                            book = load_workbook(file_path)
                            sheet_name = 'Table Produced'
                    

                            # Create a Pandas Excel writer using openpyxl
                            with pd.ExcelWriter(file_path, engine='openpyxl', mode= 'a', if_sheet_exists='replace') as writer:
                                df.to_excel(writer, index=False, sheet_name=sheet_name, startrow = 2, startcol=2)

                            wb2 = load_workbook(file_path)
                            ws = wb2[sheet_name]

                            # Adjust column widths
                            column_widths = {
                                "H": 25,  # Variable Name
                                "I": 25,  # Probecard - Cantilever
                                "C": 40,  # Probecard - Vertical
                                "D": 25,  # Fabrication1 - PCB/Board Repair
                                "E": 25,  # Fabrication2 - Test Sockets
                                "F": 35,  # Fabrication3 - Mechanical/General
                                "G": 25   # Total Value
                            }

                            for col, width in column_widths.items():
                                ws.column_dimensions[col].width = width

                            for row in ws.iter_rows(min_row=2, max_row=2 + len(df) + 1, min_col=2, max_col=2 + len(df.columns) + 1):
                                for cell in row:
                                    cell.alignment = Alignment(horizontal='center', vertical='center')

                            for col, width in column_widths.items():
                                ws.column_dimensions[col].width = width

                            # Set font style and size
                            font = Font(name='Tahoma', size=8)

                            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                                for cell in row:
                                    cell.font = font

                            for col in range(2, 2 + len(df.columns)):
                                cell = ws.cell(row=2, column=col)
                                cell.font = Font(bold=True)




                            gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

                            for row in ws.iter_rows(min_row=2, max_row=2 + len(df) + 1, min_col=3, max_col=2 + len(df.columns)):
                                for cell in row:
                                    cell.fill = gray_fill
                            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
                            if save_path:
                                wb.save(save_path)
                                wb2.save(save_path)
                                messagebox.showinfo("File Saved", f"File saved: {file_path}")
                            else:
                                messagebox.showerror("Error", "ERROR CANT SAVE")
                                
                        else:
                            messagebox.showerror("ERROR", "BSP rate column lettter not found.")
                    else:
                        messagebox.showerror("ERROR","\nNo column containing 'BSP RATE' was found.")
                        bsp_rate_column_letter = None

                else:
                    messagebox.showerror("error","No column containing 'BSP RATE' was found.")

            else:
                messagebox.showerror("Error", "BSP Rate value not found below the specified cell.")
                return 
            

    def save_file(self):
        global file_path
        if not self.file_uploaded:
            messagebox.showerror("Error", "Please upload a file before saving.")
            return
        
       
if __name__ == "__main__":
    root = ctk.CTk()
    app = CustomTkinterApp(root)
    root.mainloop()
