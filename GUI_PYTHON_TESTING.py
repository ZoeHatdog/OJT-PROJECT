import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os
from tkinter import Tk, Button, Entry, Label, filedialog, messagebox

def upload_file():
    global file_path
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path:
        file_label.config(text=f"File: {os.path.basename(file_path)}")

def convert_and_save():
    global file_path
    try:
        exchange_rate = float(exchange_rate_entry.get())
    except ValueError:
        messagebox.showerror("Input Error", "Please enter a valid exchange rate.")
        return

    if not file_path:
        messagebox.showerror("File Error", "Please upload an Excel file first.")
        return

    try:
        xls = pd.ExcelFile(file_path)
        df = pd.read_excel(file_path, sheet_name='Account Transactions', header=4)

        php_column_name = 'Debit (PHP)'  # Adjust based on your file
        usd_column_name = 'USD'
        final_rate = 1 / exchange_rate 

        df[usd_column_name] = df[php_column_name] * final_rate

        wb = openpyxl.load_workbook(file_path, keep_links=True)
        sheet = wb['Account Transactions']

        php_column_index = df.columns.get_loc(php_column_name) + 1
        php_column_letter = get_column_letter(php_column_index)
        usd_column_index = php_column_index + 1
        usd_column_letter = get_column_letter(usd_column_index)

        sheet[f'{usd_column_letter}5'] = usd_column_name

        for idx, value in enumerate(df[usd_column_name], start=6):
            cell = f'{usd_column_letter}{idx}'
            sheet[cell] = value

        output_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx *.xls")])
        if output_file_path:
            wb.save(output_file_path)
            messagebox.showinfo("Success", f"Modified Excel file saved as: {output_file_path}")

    except Exception as e:
        messagebox.showerror("Error", str(e))

# Create the main window
root = Tk()
root.title("Excel PHP to USD Converter")

# Create and place the widgets
upload_button = Button(root, text="Upload Excel File", command=upload_file)
upload_button.pack(pady=10)

file_label = Label(root, text="No file uploaded")
file_label.pack(pady=10)

exchange_rate_label = Label(root, text="Enter USD Exchange Rate:")
exchange_rate_label.pack(pady=10)

exchange_rate_entry = Entry(root)
exchange_rate_entry.pack(pady=10)

convert_button = Button(root, text="Convert and Save", command=convert_and_save)
convert_button.pack(pady=10)

# Start the main event loop
root.mainloop()
