import openpyxl.utils
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from prettytable import PrettyTable

#THIS FILE IS THE SAME AS CELL.PY BUT REMOVED UNNECESSARY CODES FOR FINALIZATION AND OPTIMAL TESTING 



input_file = 'BACKEND\Excel Files\MAY.xlsx'
workbook = openpyxl.load_workbook(input_file)

# Find the position of the cell that contains the word "Consumables"
sheet = workbook['Account Transactions']

#ROWS

target_value = 'CONSUMABLES' # MUST BE CAPITALIZED 
target_value2 = 'TOTAL CONSUMABLES'

factory_value =  'Factory Supplies'
total_factory = 'TOTAL FACTORY SUPPLIES'

start_row = None
end_row = None

BSP_RATE = 57.762


def row_finder(workbook,sheet,target_value,target_value2,start_row,end_row):
    for row in range(1, sheet.max_row + 1):
        
        cell_value = sheet[f'A{row}'].value
        if cell_value:
            cell_value = str(cell_value).strip().upper()  # Normalize by stripping whitespace and converting to uppercase
            if cell_value == target_value:
                start_row = row + 1
            elif cell_value == target_value2:
                end_row = row
        
        # Exit loop if both start_row and end_row are found
        if start_row == None and end_row == None:
            start_row = 0
            end_row = 0
        
        if start_row and end_row:
            break       
    return start_row, end_row

pin_name = "Cost of Raw Materials - Pins"
consumable_name = "Consumables"

consumable_value = row_finder(workbook,sheet,'CONSUMABLES','TOTAL CONSUMABLES',start_row,end_row)
pin_value = row_finder(workbook, sheet, 'COST OF RAW MATERIALS - PINS', 'TOTAL COST OF RAW MATERIALS - PINS', start_row, end_row)
vertical_head_value = row_finder(workbook, sheet, 'COST OF RAW MATERIALS - VERTICAL HEAD', 'TOTAL COST OF RAW MATERIALS - VERTICAL HEAD', start_row, end_row)
factory_value = row_finder(workbook, sheet,'FACTORY SUPPLIES', 'TOTAL FACTORY SUPPLIES', start_row, end_row)
freight_in1_value = row_finder(workbook, sheet, 'FREIGHT IN - CONSUMABLES AND TOOLING EXPENSE', 'TOTAL FREIGHT IN - CONSUMABLES AND TOOLING EXPENSE', start_row, end_row)
freight_in2_value = row_finder(workbook, sheet, 'FREIGHT IN - DIRECT MATERIALS', 'TOTAL FREIGHT IN - DIRECT MATERIALS', start_row, end_row)
outside_value = row_finder(workbook,sheet, 'OUTSIDE SERVICES/FABRICATION', 'TOTAL OUTSIDE SERVICES/FABRICATION', start_row, end_row)
tooling_value = row_finder(workbook,sheet, 'TOOLING EXPENSE','TOTAL TOOLING EXPENSE', start_row, end_row)

print("Consumable Rows", consumable_value)
print("Row of Total Cost of Raw Materials Pins", pin_value)
print("Row of Vertical Head", vertical_head_value)
print("Row of Factory Supply", factory_value)
print("Row of Freight 1", freight_in1_value)
print("Row of Freight 2", freight_in2_value)
print(" Row of Outside Services", outside_value)
print("Row of Tooling Expense", tooling_value)


#Columns 
product_column_letter = None                                                                                           
target_value3 = 'PRODUCT'
max_search_rows = 10


# Iterate through all rows and columns



for row in sheet.iter_rows(max_row=max_search_rows):
    for cell in row:
        if cell.value:
            cell_value = str(cell.value).strip().upper()  # Normalize by stripping whitespace and converting to uppercase
            if cell_value == target_value3:
                product_column_letter = openpyxl.utils.get_column_letter(cell.column)
                break
    
    
    if product_column_letter:
        break
    
start_row = start_row


if product_column_letter == "A":
    product_column_variable = 1
elif product_column_letter == "B":
    product_column_variable = 2
elif product_column_letter == "C":
    product_column_variable = 3
elif product_column_letter == "D":
    product_column_variable = 4
elif product_column_letter == "E":
    product_column_variable =5
elif product_column_letter == "F":
    product_column_variable = 6
elif product_column_letter == "H":
    product_column_variable = 7
elif product_column_letter == "I":
    product_column_variable = 8
elif product_column_letter == "J":
    product_column_variable = 9
column_data = []
# Extract data from column I (column index 9, assuming 'Product' is in column I)
column_i_data = []



for row_num in range(outside_value[0], outside_value[1]):
    cell_value = sheet.cell(row=row_num, column=product_column_variable+1).value  # Column I corresponds to index 9 (A=1, B=2, ..., I=9)
    column_i_data.append((row_num, cell_value)) 

# Print or process the data

print("Data from column I (Product):") 
for value in column_i_data:                  #CHECKS DATA FOR RETURNING THE VALUE PRODUCT FROM START AND END ROW
    print(value) 



#THIS CODE FINDS THE COLUMN OF DEBIT 
debit_colum_row = None
target_value4 = 'Debit (PHP)'
max_search_rows = 10



for row in sheet.iter_rows(max_row=max_search_rows): #SEARCHING FOR THE DEBIT COLUMN
    for cell in row:
        if cell.value == target_value4:
            debit_colum_row = cell
            break
    if debit_colum_row:
        break

debit_column_letter = debit_colum_row.coordinate[0] #Determines the column of the DEBIT(PHP)




print("---------------------------")


def categorize_rows(sheet, consumable_value_range, product_column_variable):
    """
    Categorize rows based on the product descriptions found in a specified column of a given sheet.

    Parameters:
    sheet (object): The sheet object to read data from.
    consumable_value_range (tuple): A tuple containing the start and end row numbers (inclusive).
    product_column_variable (int): The column index (0-based) where the product descriptions are located.

    Returns:
    dict: A dictionary with categorized rows.
    """

    # Initialize arrays to hold categorized row numbers
    fabrication1 = []
    fabrication2 = []
    fabrication3 = []
    probe_card = []
    probe_cardv = []

    # Dictionary to map product descriptions to their corresponding arrays
    product_mapping = {
        'Fabrication1 - PCB/Board Repair': fabrication1,
        'Fabrication2 - Test Sockets': fabrication2,
        'Fabrication3 - Mechanical/General': fabrication3,
        'Probecard - Vertical': probe_cardv,
        'Probecard - Cantilever': probe_card
    }

    # Iterate through the specified range of rows
    for row in range(consumable_value_range[0], consumable_value_range[1]):  # <-------------------------------------------
        cell_value = sheet.cell(row=row, column=product_column_variable + 1).value
        print(f"Row {row}, Column {product_column_variable + 1}: {cell_value}")  # Debug statement

        if cell_value is None:
            continue  # Skip None values

        # Add the row number to the corresponding array
        if cell_value in product_mapping:
            product_mapping[cell_value].append(row)

    return product_mapping



def getting_data(sheet, debit_column_letter, big_data, BSP_rate):
    fab1_final_value = 0
    probe_card_final_value = 0
    probe_cardv_final_value = 0
    fab2_final_value = 0
    fab3_final_value = 0


    def get_cell_value(sheet, cell_coordinate):
        return sheet[cell_coordinate].value
    
    probe_card = big_data.get('Probecard - Cantilever', [])
    probe_cardv = big_data.get('Probecard - Vertical', [])
    fabrication1 = big_data.get('Fabrication1 - PCB/Board Repair', [])
    fabrication2 = big_data.get('Fabrication2 - Test Sockets', [])
    fabrication3 = big_data.get('Fabrication3 - Mechanical/General', [])

    for rows in probe_card:
        cell_coordinate = f"{debit_column_letter}{rows}"
        cell_value = get_cell_value(sheet, cell_coordinate)
        if cell_value is not None:
            probe_card_final_value += cell_value
    
    for rows in probe_cardv:
        cell_coordinate = f"{debit_column_letter}{rows}"
        cell_value = get_cell_value(sheet, cell_coordinate)
        if cell_value is not None:
            probe_cardv_final_value += cell_value

    for rows in fabrication1:
        cell_coordinate = f"{debit_column_letter}{rows}"
        cell_value = get_cell_value(sheet, cell_coordinate)
        if cell_value is not None:
            fab1_final_value += cell_value

    for rows in fabrication2:
        cell_coordinate = f"{debit_column_letter}{rows}"
        cell_value = get_cell_value(sheet, cell_coordinate)
        if cell_value is not None:
            fab2_final_value += cell_value

    for rows in fabrication3:
        cell_coordinate = f"{debit_column_letter}{rows}"
        cell_value = get_cell_value(sheet, cell_coordinate)
        if cell_value is not None:
            fab3_final_value += cell_value

    # Convert to float and format the final values
    probe_card_final_value = float(probe_card_final_value)
    probe_cardv_final_value = float(probe_cardv_final_value)
    fab1_final_value = float(fab1_final_value)
    fab2_final_value = float(fab2_final_value)
    fab3_final_value = float(fab3_final_value)

    probe_card_final_value = "{:.2f}".format(probe_card_final_value / BSP_RATE)
    probe_cardv_final_value = "{:.2f}".format(probe_cardv_final_value / BSP_RATE)
    fab1_final_value = "{:.2f}".format(fab1_final_value / BSP_RATE)
    fab2_final_value = "{:.2f}".format(fab2_final_value / BSP_RATE)
    fab3_final_value = "{:.2f}".format(fab3_final_value / BSP_RATE)

    # Calculate total value
    total_value = float(probe_card_final_value) + float(probe_cardv_final_value) + float(fab1_final_value) + float(fab2_final_value) + float(fab3_final_value)
    total_value = "{:.2f}".format(total_value)

    return probe_card_final_value, probe_cardv_final_value, fab1_final_value, fab2_final_value, fab3_final_value, total_value

consume_row = categorize_rows(sheet, consumable_value, product_column_variable)
consume_data = getting_data(sheet, debit_column_letter, consume_row, BSP_RATE)

pin_row = categorize_rows(sheet, pin_value, product_column_variable)
pin_data = getting_data(sheet, debit_column_letter, pin_row, BSP_RATE)

vertical_row = categorize_rows(sheet, vertical_head_value, product_column_variable)
vertical_data = getting_data(sheet, debit_column_letter, vertical_row, BSP_RATE)

factory_row = categorize_rows(sheet, factory_value, product_column_variable)
factory_data = getting_data(sheet, debit_column_letter, factory_row, BSP_RATE)

freight1_row = categorize_rows(sheet, freight_in1_value, product_column_variable)
freight1_data = getting_data(sheet, debit_column_letter, freight1_row, BSP_RATE)

freight2_row = categorize_rows(sheet, freight_in2_value, product_column_variable)
freight2_data = getting_data(sheet, debit_column_letter, freight2_row, BSP_RATE)

outside_row = categorize_rows(sheet, outside_value, product_column_variable)
outside_data = getting_data(sheet, debit_column_letter, outside_row, BSP_RATE)

tooling_row = categorize_rows(sheet, tooling_value, product_column_variable)
tooling_data = getting_data(sheet, debit_column_letter, tooling_row, BSP_RATE)

def replace_zero_values(data):
    return ["-" if value == "0.00" else value for value in data]


data_variables = [
    ("Consumables", consume_data),
    ("Cost of Raw Materials - Pins", pin_data),
    ("Cost of Raw Materials - Vertical head", vertical_data),
    ("Factory Supplies", factory_data),
    ("Freight In - Consumables And Tooling Expense", freight1_data),
    ("Freight In - Direct Materials", freight2_data),
    ("Outside Services/Fabrication", outside_data),
    ("Tooling Expense", tooling_data)
]


#ADDITIONAL OF TOTAL VALUES IN COLUMN
probe_card_value = 0
probe_card_value = float(consume_data[0]) + float(pin_data[0]) + float(vertical_data[0]) + float(factory_data [0]) + float(freight1_data[0]) + float(freight2_data[0]) + float(outside_data[0]) + float(tooling_data[0])
probe_cardv_value = float(consume_data[1]) + float(pin_data[1]) + float(vertical_data[1]) + float(factory_data [1]) + float(freight1_data[1]) + float(freight2_data[1]) + float(outside_data[1]) + float(tooling_data[1])
fabrication1_value = float(consume_data[2]) + float(pin_data[2]) + float(vertical_data[2]) + float(factory_data [2]) + float(freight1_data[2]) + float(freight2_data[2]) + float(outside_data[2]) + float(tooling_data[2])
fabrication2_value = float(consume_data[3]) + float(pin_data[3]) + float(vertical_data[3]) + float(factory_data [3]) + float(freight1_data[3]) + float(freight2_data[3]) + float(outside_data[3]) + float(tooling_data[3])
fabrication3_value = float(consume_data[4]) + float(pin_data[4]) + float(vertical_data[4]) + float(factory_data [4]) + float(freight1_data[4]) + float(freight2_data[4]) + float(outside_data[4]) + float(tooling_data[4])
total_total_value = float(consume_data[5]) + float(pin_data[5]) + float(vertical_data[5]) + float(factory_data [5]) + float(freight1_data[5]) + float(freight2_data[5]) + float(outside_data[5]) + float(tooling_data[5])
print("probe value is:", probe_card_value)
print("2", probe_cardv_value)
print("3", fabrication1_value)
print("4", fabrication2_value)
print("5", fabrication3_value)
print("6", total_total_value)

print("Consume Data: DATA",consume_data[1])
# Extract the variable names and data

variable_names = [var[0] for var in data_variables]
data_values = [replace_zero_values(var[1]) for var in data_variables]

# Create a DataFrame
df = pd.DataFrame(data_values, columns=[
    "Probecard - Cantilever",
    "Probecard - Vertical",
    "Fabrication1 - PCB/Board Repair",
    "Fabrication2 - Test Sockets",
    "Fabrication3 - Mechanical/General",
    "Total Value"
])


# Add the variable names as a column
df.insert(0, "Variable Name", variable_names)

new_row = {'Total Value',probe_card_value, probe_cardv_value, fabrication1_value,fabrication2_value,fabrication3_value,total_total_value}
new_row = {
    "Variable Name": 'Total Value',  # Assuming 'Total Value' is the name for the first column
    "Probecard - Cantilever": probe_card_value,
    "Probecard - Vertical": probe_cardv_value,
    "Fabrication1 - PCB/Board Repair": fabrication1_value,
    "Fabrication2 - Test Sockets": fabrication2_value,
    "Fabrication3 - Mechanical/General": fabrication3_value,
    "Total Value": total_total_value
}

# Append the new row to the DataFrame
new_row_df = pd.DataFrame([new_row])

# Concatenate the original DataFrame with the new row DataFrame
df = pd.concat([df, new_row_df], ignore_index=True)



# Concatenate the original DataFrame with the new row DataFrame

book = load_workbook(input_file)
sheet_name = 'Table Produced'


# Create a Pandas Excel writer using openpyxl
with pd.ExcelWriter(input_file, engine='openpyxl', mode= 'a', if_sheet_exists='replace') as writer:
    df.to_excel(writer, index=False, sheet_name=sheet_name, startrow = 2, startcol=2)

wb = load_workbook(input_file)
ws = wb[sheet_name]

# Adjust column widths
column_widths = {
    "H": 25,  # Variable Name
    "I": 25,  # Probecard - Cantilever
    "C": 40,  # Probecard - Vertical
    "D": 25,  # Fabrication1 - PCB/Board Repair
    "E": 25,  # Fabrication2 - Test Sockets
    "F": 35,  # Fabrication3 - Mechanical/General
    "G": 25   # Total Value
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

for row in ws.iter_rows(min_row=2, max_row=2 + len(df) + 1, min_col=2, max_col=2 + len(df.columns) + 1):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Set font style and size
font = Font(name='Tahoma', size=8)

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.font = font

for col in range(2, 2 + len(df.columns)):
    cell = ws.cell(row=2, column=col)
    cell.font = Font(bold=True)




gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

for row in ws.iter_rows(min_row=2, max_row=2 + len(df) + 1, min_col=3, max_col=2 + len(df.columns)):
    for cell in row:
        cell.fill = gray_fill





# Save the workbook with the adjusted column widths
wb.save(input_file)
# Display a message indicating success
print(f"The data has been successfully written to sheet '{sheet_name}' in {input_file}")