import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os




# Load the provided Excel file using pandas


"""
file_path = 'BACKEND\\excel_test.xlsx'  # Replace with the actual file path if needed
xls = pd.ExcelFile(file_path)

# Load the data from the 'Account Transactions' sheet
df = pd.read_excel(file_path, sheet_name='Account Transactions', header=4)  # header=4 to start reading from Row 5

# Print the column names to identify the exact name of the PHP column
print("Column names in the dataframe:")
for i, col in enumerate(df.columns):
    print(f"{i}: '{col}'")

# After identifying the exact column name, use it in the following code
# For this example, let's assume the correct column name is 'PHP'

php_column_name = 'Debit (PHP)'  # Replace with the actual column name after inspection
usd_column_name = 'USD'

# Define the exchange rate from PHP to USD (e.g., 1 PHP = 0.018 USD)
usdollar = 56.911
exchange_rate = 1 / usdollar
print("Exchange Rate = " + str(exchange_rate))

# Convert the PHP values to USD and create a new column
df[usd_column_name] = df[php_column_name] * exchange_rate

# Display the first few rows of the modified dataframe to verify
print(df.head())

# Load the original workbook using openpyxl to retain formatting
wb = openpyxl.load_workbook(file_path, keep_links=True)
sheet = wb['Account Transactions']

# Identify the column letter for the PHP column
php_column_index = df.columns.get_loc(php_column_name) + 1
php_column_letter = get_column_letter(php_column_index)

# Determine the column letter for the new USD column
usd_column_index = php_column_index + 1
usd_column_letter = get_column_letter(usd_column_index)

# Write the USD column header in the appropriate place
sheet[f'{usd_column_letter}5'] = usd_column_name  # Assuming the headers are on row 5

# Write the USD values starting from the appropriate row
for idx, value in enumerate(df[usd_column_name], start=6):  # start=6 because data starts from row 6
    cell = f'{usd_column_letter}{idx}'
    sheet[cell] = value

# Define the output file path and save the modified workbook
output_file_path = 'BACKEND\\modified_excel_test_2.xlsx'

# Check if the output file already exists and delete it if it does to avoid PermissionError
if os.path.exists(output_file_path):
    os.remove(output_file_path)

wb.save(output_file_path)
print(f"Modified Excel file saved as: {output_file_path}")
"""

#Read product description and input number and create Table
'BACKEND\\excel_test.xlsx' 

file_path = 'BACKEND\\modified_excel_test.xlsx'
xls = pd.ExcelFile(file_path)

df = pd.read_excel(file_path, sheet_name='Account Transactions', header=4) #Header 4 is the start of the reading of the file starting from the row #4 

print("Column names in the dataframe:")
for i, col in enumerate(df.columns):
    print(f"{i}: '{col}'")

# Assume the correct column name is 'Product' after inspection
product_column_name = 'Product'  # Replace with the actual column name after inspection

# Extract all the strings from the "Product" column and remove repetitions
unique_products = df[product_column_name].dropna().unique()

# Display the unique products (Extract "Unique" Data from the product column )
print("Unique products:")
for product in unique_products:
    print(product)


# Create a new DataFrame with unique products as columns
unique_products_df = pd.DataFrame(columns=unique_products)

# Save the new DataFrame to a new Excel file
output_file_path = 'BACKEND\\unique_products_table.xlsx'
unique_products_df.to_excel(output_file_path, index=False)

print(f"Table with unique products as columns saved to: {output_file_path}")