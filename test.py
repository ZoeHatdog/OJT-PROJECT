import pandas as pd

# Replace 'file_path.xlsx' with the path to your Excel file
file_path = 'BACKEND\excel_test.xlsx'

# Read the Excel file
df = pd.read_excel(file_path)

# Find the row with the word 'total'
total_row = df[df.apply(lambda row: row.astype(str).str.contains('Total Cost of Raw Materials - Pins', case=False).any(), axis=1)]


# Locate Specific Number 
if not total_row.empty:
    total_row_index = total_row.index[0]
    total_row_values = df.iloc[total_row_index]

    # Find the number nearest to 'total' on the same row
    nearest_number = None
    for value in total_row_values:
        if isinstance(value, (int, float)) and not pd.isna(value):
            nearest_number = value
            break

    if nearest_number is not None:
        print("Number nearest to 'total':", nearest_number)
    else:
        print("No number found near 'total' in the same row.")
else:
    print("Word 'total' not found in the Excel file.")

#Locating BSP rate 

   

    # Load the provided Excel file
file_path = 'BACKEND\excel_test.xlsx'  # Replace with the actual file path if needed
xls = pd.ExcelFile(file_path)

    # Load the data from the 'Account Transactions' sheet
df = pd.read_excel(file_path, sheet_name='Account Transactions')

    # Find the cell with "BSP RATE" (assuming it is "(BSP Rate - Average Monthly)")
bsp_rate_cell = df[df.isin(["(BSP Rate - Average Monthly)"]).any(axis=1)]

    # Get the index of the row containing "BSP RATE"
bsp_rate_row_index = bsp_rate_cell.index[0]

    # Get the column index of "BSP RATE"
bsp_rate_col_index = bsp_rate_cell.columns[bsp_rate_cell.iloc[0] == "(BSP Rate - Average Monthly)"][0]

    # Find the nearest number below "BSP RATE" in the same column
for i in range(bsp_rate_row_index + 1, len(df)):
        value = df.iloc[i, df.columns.get_loc(bsp_rate_col_index)]
        if pd.notna(value) and isinstance(value, (int, float)):
            bsp_rate_value = value
            break

print(f"BSP RATE value: {bsp_rate_value}")

#READING THE PHP COLUMN AND ASSIGNING TO A VARIABLE 
"""
file_path = 'BACKEND\excel_test.xlsx'   # Replace with the actual file path if needed
xls = pd.ExcelFile(file_path)

    # Load the data from the 'Account Transactions' sheet
df = pd.read_excel(file_path, sheet_name='Account Transactions')

# Print the column names to identify the exact name of the PHP column
print("Column names in the dataframe:")
for i, col in enumerate(df.columns):
    print(f"{i}: '{col}'")

# Now, let's assume the correct column name after inspecting
php_column_name = 'Debit (PHP)'  # Replace with the actual column name after inspection

# Extract all numeric values from the "Debit (PHP)" column and assign them to a list
php_values = []
for value in df[php_column_name]:
    if pd.notna(value) and isinstance(value, (int, float)):
        php_values.append(value)

# Print all numbers
for number in php_values:
    print(number)

# Assign the list to a variable
php_numbers = php_values
"""
#PRINT COLUMN NAMES 
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os

# Load the provided Excel file using pandas
file_path = 'BACKEND\\excel_test.xlsx'  # Replace with the actual file path if needed
xls = pd.ExcelFile(file_path)

# Load the data from the 'Account Transactions' sheet
df = pd.read_excel(file_path, sheet_name='Account Transactions', header=4)  # header=4 to start reading from Row 5

# Print the column names to identify the exact name of the PHP column
print("Column names in the dataframe:")
for i, col in enumerate(df.columns):
    print(f"{i}: '{col}'")

# After identifying the exact column name, use it in the following code
# For this example, let's assume the correct column name is 'PHP'

php_column_name = 'Debit (PHP)'  # Replace with the actual column name after inspection
usd_column_name = 'USD'

# Define the exchange rate from PHP to USD (e.g., 1 PHP = 0.018 USD)
usdollar = 56.911
exchange_rate = 1 / usdollar
print("Exchange Rate = " + str(exchange_rate))

# Convert the PHP values to USD and create a new column
df[usd_column_name] = df[php_column_name] * exchange_rate

# Display the first few rows of the modified dataframe to verify
print(df.head())

# Load the original workbook using openpyxl to retain formatting
wb = openpyxl.load_workbook(file_path, keep_links=True)
sheet = wb['Account Transactions']

# Identify the column letter for the PHP column
php_column_index = df.columns.get_loc(php_column_name) + 1
php_column_letter = get_column_letter(php_column_index)

# Determine the column letter for the new USD column
usd_column_index = php_column_index + 1
usd_column_letter = get_column_letter(usd_column_index)

# Write the USD column header in the appropriate place
sheet[f'{usd_column_letter}5'] = usd_column_name  # Assuming the headers are on row 5

# Write the USD values starting from the appropriate row
for idx, value in enumerate(df[usd_column_name], start=6):  # start=6 because data starts from row 6
    cell = f'{usd_column_letter}{idx}'
    sheet[cell] = value

# Define the output file path and save the modified workbook
output_file_path = 'BACKEND\\modified_excel_test.xlsx'

# Check if the output file already exists and delete it if it does to avoid PermissionError
if os.path.exists(output_file_path):
    os.remove(output_file_path)

wb.save(output_file_path)
print(f"Modified Excel file saved as: {output_file_path}")

"""
file_path = 'BACKEND\excel_test.xlsx'   # Replace with the actual file path if needed
xls = pd.ExcelFile(file_path)


# Load the data from the 'Account Transactions' sheet
df = pd.read_excel(file_path, sheet_name='Account Transactions', header=4)  # header=4 to start reading from Row 5

# Print the column names to identify the exact name of the PHP column
print("Column names in the dataframe:")
for i, col in enumerate(df.columns):
    print(f"{i}: '{col}'")

# After identifying the exact column name, use it in the following code
# For this example, let's assume the correct column name is 'Debit (PHP)'

php_column_name = 'Debit (PHP)'  # Replace with the actual column name after inspection
usd_column_name = 'Debit (USD)'

# Define the exchange rate from PHP to USD (e.g., 1 PHP = 0.018 USD)
usdollar = 56.911
exchange_rate = 1 / usdollar
print("Exchange Rate = " + str(exchange_rate))


# Convert the PHP values to USD and create a new column
df[usd_column_name] = df[php_column_name] * exchange_rate

# Display the first few rows of the modified dataframe to verify
print(df.head())

# Save the modified dataframe to a new Excel file
output_file_path = 'BACKEND\modified_excel_test.xlsx'  # Replace with the desired output file path
df.to_excel(output_file_path, index=False)
"""


